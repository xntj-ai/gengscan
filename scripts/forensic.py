# -*- coding: utf-8 -*-
"""Column-level numerical forensic detectors for spreadsheet (.xlsx) data.

Scans every column (and column pair) of each sheet for arithmetic signals that
cannot plausibly be explained by honest measurement error. Detector families:

  A  Constant inter-column difference  — two columns differ by a fixed constant
  B  Terminal-digit lock              — last digit is almost always the same value
  C  Arithmetic progression in a column — constant first-difference run
  D  Exact-value domination           — one value repeats far too often
  E  Last-two-decimals repetition     — many values share the same last 2 decimals
  F  Duplicate blocks                  — identical numeric sequence across columns
                                          (copy-paste fingerprint), within/across sheets
  G  Block-level decimal lock          — two sheets share decimal parts but differ in
                                          integer parts ("kept decimals, changed integers")
  H  Row-level duplicate blocks        — a measurement row copied to another sample/species

Outputs a short "candidate, needs manual review" list. An anomaly is NOT proof of
fabrication; it is a lead for a human to inspect the measurement columns.

Axis/independent-variable columns (scan axes, step grids) are excluded to reduce
false positives. An "omics gate" suppresses bulk row-duplication from omics tables
(abundance matrices, radiomics features, phylogenetic trees, sequencing read IDs).

Usage:
    python forensic.py [file1.xlsx file2.xlsx ...]
    python forensic.py --indir ./papers --out report.txt

With no file arguments, scans ./papers/*/*.xlsx by default.
"""
import sys, glob, math, argparse, os
from collections import Counter
import openpyxl


def ndec(nf):
    if not nf or nf in ('General', '@'):
        return None
    nf = nf.split(';')[0]
    if '.' in nf:
        n = 0
        for ch in nf.split('.')[1]:
            if ch in '0#':
                n += 1
            else:
                break
        return n
    return 0


def eff_ndec(v, nf):
    """有效小数位:优先单元格格式;General 时从数值本身推断(防止漏检 General 格式列)。"""
    nd = ndec(nf)
    if nd is not None:
        return nd
    s = repr(float(v))
    if 'e' in s or 'E' in s:
        return None
    return len(s.split('.')[1]) if '.' in s else 0


def sheet_columns(ws, max_rows=4000):
    """返回 {col_index: [(value, ndecimals), ...]} ,按出现行顺序。"""
    cols = {}
    for ri, row in enumerate(ws.iter_rows()):
        if ri > max_rows:
            break
        for ci, c in enumerate(row):
            v = c.value
            if isinstance(v, (int, float)) and not isinstance(v, bool) and math.isfinite(v):
                cols.setdefault(ci, []).append((ri, float(v), eff_ndec(v, c.number_format)))
    return cols


def has_real_fraction(vals):
    """该列是否真有小数(排除整数被显示成 x.0 的情形)。"""
    return any(abs(v - round(v)) > 1e-9 for v in vals)


def is_axis(series):
    """判断是否为自变量/坐标轴列(应从造假检测中排除):
       - 整列单调(扫描轴),或
       - 等步长扫描(一阶差几乎恒定且覆盖>=85%),或
       - 小数部分只有极少数取值(步进轴)。"""
    vals = [v for _, v, _ in series]
    n = len(vals)
    if n < 6:
        return False
    diffs = [vals[i+1] - vals[i] for i in range(n-1)]
    pos = sum(1 for d in diffs if d > 1e-12)
    neg = sum(1 for d in diffs if d < -1e-12)
    if pos >= 0.97 * (n-1) or neg >= 0.97 * (n-1):   # 严格单调 = 轴
        return True
    if diffs:
        d0 = diffs[0]
        if abs(d0) > 1e-12 and sum(1 for d in diffs if abs(d-d0) < 1e-9) >= 0.85 * (n-1):
            return True                              # 等步长扫描
    # 注:不再用"小数部分单一"判轴——那会误杀"末两位全是44"这类造假列。
    return False


def chk_terminal(series):
    """B 单列末位锁定 + E 小数末两位雷同。"""
    flags = []
    vals = [v for _, v, _ in series]
    disp = []
    for _, v, nd in series:
        if nd and nd >= 1:
            disp.append(f'{v:.{nd}f}')
    if len(disp) >= 20 and has_real_fraction(vals):
        last = [s[-1] for s in disp]
        c = Counter(last)
        top, topn = c.most_common(1)[0]
        # 只报"锁定到非0数字"——尾零=显示精度>数据精度的假象,排除
        if top != '0' and topn / len(last) >= 0.85:
            flags.append(('B 末位锁定', f'{len(last)}个值末位{topn}个是"{top}"({topn*100//len(last)}%)'))
    # 小数末两位雷同(同样排除 "00" 尾零假象)
    d2 = [f'{v:.{nd}f}'[-2:] for _, v, nd in series if nd and nd >= 2]
    if len(d2) >= 15 and has_real_fraction(vals):
        c2 = Counter(d2)
        t2, t2n = c2.most_common(1)[0]
        if t2 != '00' and t2n / len(d2) >= 0.7:
            flags.append(('E 小数末两位雷同', f'{len(d2)}个值有{t2n}个末两位="{t2}"'))
    return flags


def chk_arith(series):
    """C 列内等差:连续一阶差为非零常数,长度>=6。"""
    vals = [v for _, v, _ in series]
    if len(vals) < 6 or not has_real_fraction(vals):
        return []
    best = 0
    i = 0
    n = len(vals)
    while i < n - 1:
        d = round(vals[i+1] - vals[i], 9)
        if d == 0:
            i += 1
            continue
        j = i
        while j < n - 1 and abs((vals[j+1] - vals[j]) - d) < 1e-9:
            j += 1
        run = j - i + 1
        if run > best:
            best = run; bd = d
        i = j if j > i else i + 1
    # 只报覆盖整列绝大部分的等差(局部线性段是平滑曲线常态,不算)
    if best >= max(10, int(0.8 * len(vals))):
        return [('C 整列等差', f'{len(vals)}个值中连续{best}个等差(公差{bd})')]
    return []


def chk_dup(series):
    """D 精确重复:连续量里同一值占比过高 / 单值碾压填充。"""
    vals = [v for _, v, _ in series]
    if len(vals) < 12 or not has_real_fraction(vals):
        return []
    c = Counter(round(v, 9) for v in vals)
    mc = c.most_common(2)
    val, cnt = mc[0]
    runner = mc[1][1] if len(mc) > 1 else 0
    distinct = len(c)
    if abs(val) <= 1e-9:
        return []
    # D 同值霸占:占比过半。注:"单值碾压填充"(低占比单值大量重复)曾尝试降阈值,
    # 但干净论文的高精度计算值(单值碾压几十倍)同形→假阳性,无法区分,
    # 故 constant-fill 归入不可靠自动检测类(软信号),不降阈值。
    if cnt >= max(6, 0.5 * len(vals)) and distinct >= 4:
        return [('D 同值霸占', f'值{val}在{len(vals)}个中出现{cnt}次')]
    return []


def chk_const_diff(cols):
    """A 两列恒定差:对齐行后两列差为非零常数。"""
    flags = []
    keys = sorted(cols)
    series_by_row = {}
    for ci in keys:
        series_by_row[ci] = {ri: v for ri, v, _ in cols[ci]}
    for a in range(len(keys)):
        for b in range(a + 1, len(keys)):
            ca, cb = keys[a], keys[b]
            common = set(series_by_row[ca]) & set(series_by_row[cb])
            if len(common) < 8:
                continue
            diffs = [round(series_by_row[ca][r] - series_by_row[cb][r], 9) for r in common]
            d0 = diffs[0]
            if abs(d0) < 1e-9:
                continue
            if all(abs(d - d0) < 1e-9 for d in diffs):
                # 排除两列本身都是常数列
                if len(set(round(series_by_row[ca][r], 9) for r in common)) > 1:
                    flags.append(('A 两列恒差', f'第{ca+1}列 - 第{cb+1}列 = {d0}(共{len(common)}行)'))
    return flags


def _is_shared_axis_or_baseline(vals):
    """该序列是否是"绘图共享轴"或"基线零列"——F 重复块的两大良性来源:
       ① 单调(共享 x 轴:波长/电压/时间,每条曲线配同一列) → 跨列相同=正常
       ② 大面积为 0(基线/虚部起点) → 跨列相同=正常
       排除这两类后,F 才聚焦"两条本应独立的测量序列雷同"。"""
    n = len(vals)
    if n < 6:
        return False
    diffs = [vals[i+1] - vals[i] for i in range(n-1)]
    pos = sum(1 for d in diffs if d > 1e-12)
    neg = sum(1 for d in diffs if d < -1e-12)
    if pos >= 0.95 * (n-1) or neg >= 0.95 * (n-1):      # 单调=共享轴
        return True
    if sum(1 for v in vals if abs(v) < 1e-9) >= 0.4 * n:  # ≥40% 为 0=基线列
        return True
    return False


def chk_dup_blocks(allcols, minlen=6):
    """F 重复块:不同列(同sheet/跨sheet/跨文件)出现相同数值序列=copy-paste造假指纹。
       精确 + 四舍五入(取整/1位小数)三档匹配,排除整列近常数 + 共享轴/基线列。"""
    flags = []
    seen_pairs = set()
    # 精确匹配:短序列即可疑;四舍五入匹配:要求更长(短列取整后易偶然碰撞)
    for nd, lab, ml in [(8, '精确', 8), (0, '取整后', 20), (1, '1位小数后', 20)]:
        buckets = {}
        for fn, sn, ci, vals in allcols:
            if len(vals) < max(minlen, ml):
                continue
            r = tuple(round(v, nd) for v in vals)
            if len(set(r)) < 4:                         # 整列近常数/取值过少,跳过
                continue
            if _is_shared_axis_or_baseline(vals):       # 共享轴/基线列=良性复用,跳过
                continue
            buckets.setdefault(r, []).append((fn, sn, ci))
        for key, locs in buckets.items():
            if len(locs) < 2:
                continue
            pk = tuple(sorted(locs))
            if pk in seen_pairs:
                continue
            seen_pairs.add(pk)
            where = '; '.join(f'{f}/{s}列{c+1}' for f, s, c in locs[:6])
            same_sheet = len(set((f, s) for f, s, c in locs)) == 1
            # 同一张图内两列相同=强嫌疑;跨图相同=常为main+SI正常复用=弱
            tag = 'F 同图内重复块' if same_sheet else 'F~ 跨图重复(可能正常复用)'
            flags.append({'file': locs[0][0], 'sheet': '(重复块)', 'tag': tag,
                          'msg': f'{len(key)}个值序列在{len(locs)}列{lab}相同: {where}'})
    return flags


def _block_fracmap(ws, nd=2, max_rows=400, max_cells=400):
    """收集一个 sheet 含真实小数的数值格,按(行,列)定位:{(ri,ci):(frac2,int)}。
       「保留小数改整数」只发生在手工小结果表,超大 sheet(>max_rows/>max_cells)
       直接放弃——既合方法论又避免在海量行原始数据上 O(n²) 卡死。None=跳过。"""
    m = {}
    for ri, row in enumerate(ws.iter_rows()):
        if ri >= max_rows:
            return None
        for ci, c in enumerate(row):
            v = c.value
            if isinstance(v, (int, float)) and not isinstance(v, bool) and math.isfinite(v):
                if abs(v - round(v)) > 1e-9:
                    m[(ri, ci)] = (round((v - math.floor(v)) * (10 ** nd)), math.floor(v))
                    if len(m) > max_cells:
                        return None
    return m


def chk_frac_block_lock(blocks, min_cells=20, frac_thresh=0.9, int_thresh=0.2):
    """G 块级跨表小数锁定(「保留小数、改整数」伪造):
       两个 sheet 逐格对齐后小数部分大面积相同、整数部分明显不同
       → 真实数据被复用、只改整数位伪造另一组。
       blocks: [(file, sheet, {(ri,ci):(frac,int)})]。"""
    flags = []
    for i in range(len(blocks)):
        for j in range(i + 1, len(blocks)):
            fa, sa, ma = blocks[i]
            fb, sb, mb = blocks[j]
            common = set(ma) & set(mb)
            if len(common) < min_cells:
                continue
            matched_fracs = [ma[k][0] for k in common if ma[k][0] == mb[k][0]]
            frac_same = len(matched_fracs)
            int_diff = sum(1 for k in common if ma[k][1] != mb[k][1])
            if not (frac_same / len(common) >= frac_thresh and int_diff / len(common) >= int_thresh):
                continue
            # 防粗网格错位假阳性(全是.0/.5的光谱数据两列错位→小数天然重合):
            # 要求匹配上的小数值【足够多样】,否则只是少数小数取值的偶然碰撞。
            fc = Counter(matched_fracs)
            distinct = len(fc)
            dom_ratio = fc.most_common(1)[0][1] / frac_same
            if distinct < 6 or dom_ratio > 0.6:
                continue
            same_file = fa == fb
            tag = 'G 块级小数锁定' if same_file else 'G~ 跨文件小数锁定'
            flags.append({'file': fa, 'sheet': f'{sa}~{sb}', 'tag': tag,
                          'msg': f'{len(common)}格对齐 小数相同{frac_same}'
                                 f'({frac_same*100//len(common)}%,{distinct}种) 整数不同{int_diff}'
                                 f'({int_diff*100//len(common)}%)'})
    return flags


def chk_dup_rows(rows, minlen=5):
    """H 行级重复块(整行数值序列被复制到另一个样本/物种行)。
       其余检测都是列向,看不到"把一行测量复制到另一行"的造假。
       rows: [(file, sheet, ridx, label, (v1,v2,...))]。同 sheet 跨行精确相同=强嫌疑。
       排除整行近常数/取值过少(避免全0行、单值行误报)。"""
    flags = []
    buckets = {}
    for fn, sn, ridx, label, vals in rows:
        if len(vals) < minlen:
            continue
        r = tuple(round(v, 6) for v in vals)
        if len(set(r)) < 4:                          # 行内取值过少(近常数)跳过
            continue
        if _is_shared_axis_or_baseline(list(vals)):  # 单调=共享浓度/x梯度行(每样本复用,良性)
            continue
        buckets.setdefault((fn, sn, r), []).append((ridx, label))
    for (fn, sn, r), locs in buckets.items():
        # 组=2~3:测量行被复制到另一两处=copy-paste嫌疑;组>=6=共享模板排除。
        if not (2 <= len(locs) <= 3):
            continue
        # 关键:必须跨>=2个不同的非空样本标签。
        # 同序列出现在无标签/同标签相邻行=副本/重复测量存储=良性。
        labels = {str(l).strip() for _, l in locs if isinstance(l, str) and l.strip()}
        if len(labels) < 2:
            continue
        where = ' / '.join(f'行{ri}({lb})' for ri, lb in locs[:6])
        flags.append({'file': fn, 'sheet': sn, 'tag': 'H 行级重复块',
                      'msg': f'{len(r)}值整行跨{len(labels)}个不同样本相同: {where}'})
    # 组学/计算表闸:真 copy-paste 是零星几行;
    # 单 sheet 出现大量行级重复 = 组学批量产出(微生物组丰度/radiomics 特征/系统发育树/
    # 测序 readID/环状RNA/基因符号),行雷同是数据本质非造假。按 sheet 计数超阈值整体抑制。
    OMICS_CAP = 8
    per_sheet = Counter((f['file'], f['sheet']) for f in flags)
    flags = [f for f in flags if per_sheet[(f['file'], f['sheet'])] <= OMICS_CAP]
    return flags


def audit_paths(paths):
    """对多个 xlsx 跑全部检查,返回结构化候选 [{file,sheet,tag,msg}],供批量审计调用。"""
    import openpyxl as _ox
    flags = []
    allcols = []
    blocks = []
    allrows = []
    for path in paths:
        try:
            wb = _ox.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        fn = path.replace('\\', '/').split('/')[-1]
        for ws in wb.worksheets:
            cols = sheet_columns(ws)
            meas = {ci: s for ci, s in cols.items() if len(s) >= 6 and not is_axis(s)}
            for ci, series in meas.items():
                allcols.append((fn, ws.title, ci, [v for _, v, _ in series]))
                for tag, msg in chk_terminal(series) + chk_arith(series) + chk_dup(series):
                    flags.append({'file': fn, 'sheet': ws.title, 'tag': tag, 'msg': f'[第{ci+1}列] {msg}'})
            for tag, msg in chk_const_diff(meas):
                flags.append({'file': fn, 'sheet': ws.title, 'tag': tag, 'msg': msg})
            fm = _block_fracmap(ws)
            if fm is not None and len(fm) >= 20:
                blocks.append((fn, ws.title, fm))
            # 行级:每行数值序列(带行标签),供 chk_dup_rows
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                if ri > 4000:
                    break
                label = next((c for c in row if isinstance(c, str) and c.strip()), '')
                nums = tuple(float(c) for c in row if isinstance(c, (int, float)) and not isinstance(c, bool) and math.isfinite(c))
                if len(nums) >= 5:
                    allrows.append((fn, ws.title, ri, label, nums))
        wb.close()
    flags += chk_dup_blocks(allcols)
    flags += chk_frac_block_lock(blocks)
    flags += chk_dup_rows(allrows)
    return flags


def analyze(path, out):
    name = path.replace('\\', '/').split('/')[-1]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    hits = []
    for ws in wb.worksheets:
        cols = sheet_columns(ws)
        # 排除自变量/坐标轴列
        meas = {ci: s for ci, s in cols.items() if len(s) >= 6 and not is_axis(s)}
        sheet_flags = []
        for ci, series in meas.items():
            for tag, msg in chk_terminal(series) + chk_arith(series) + chk_dup(series):
                sheet_flags.append((tag, f'[第{ci+1}列] {msg}'))
        for tag, msg in chk_const_diff(meas):
            sheet_flags.append((tag, msg))
        if sheet_flags:
            hits.append((ws.title, sheet_flags, len(cols) - len(meas)))
    wb.close()
    out.write(f'\n### {name}\n')
    if not hits:
        out.write('  [OK] 排除坐标轴列后,无列级算术异常候选\n')
    for sn, fl, nax in hits:
        out.write(f'  * {sn}  (已排除{nax}个坐标轴/自变量列)\n')
        for tag, msg in fl:
            out.write(f'      [{tag}] {msg}\n')
    return hits


def main():
    ap = argparse.ArgumentParser(description='Column-level numerical forensic detectors for .xlsx data.')
    ap.add_argument('files', nargs='*', help='xlsx files to scan (default: ./papers/*/*.xlsx)')
    ap.add_argument('--indir', default='./papers', help='input dir scanned when no files given')
    ap.add_argument('--out', default='./forensic_report.txt', help='output report path')
    args = ap.parse_args()

    # Accept both layouts: xlsx directly in --indir, or one-subdir-per-paper.
    files = args.files or sorted(set(
        glob.glob(os.path.join(args.indir, '*.xlsx')) +
        glob.glob(os.path.join(args.indir, '*', '*.xlsx'))))
    out = open(args.out, 'w', encoding='utf-8')
    out.write('列级取证候选 — 已排除自变量/坐标轴列。异常≠造假,仅供人工核对measurement列。\n')
    total = 0
    for f in files:
        try:
            h = analyze(f, out)
            total += sum(len(x[1]) for x in h)
        except Exception as e:
            out.write(f'\n### {f}\n  ERROR {e!r}\n')
    out.write(f'\n候选总数(排除坐标轴后): {total}\n')
    out.close()
    print(f'written {args.out} | candidates {total}')


if __name__ == '__main__':
    main()
