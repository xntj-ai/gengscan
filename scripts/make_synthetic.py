# -*- coding: utf-8 -*-
"""Build a synthetic 'positive control' xlsx with planted fabrication signals.

Each sheet embeds one of the numerical-forensic detector patterns so you can verify
that forensic.py lights up the corresponding family. Includes one negative-control
sheet (normal noisy measurement) that must NOT trigger any detector.

Usage:
    python make_synthetic.py [--out ./synthetic_positive.xlsx]
"""
import openpyxl, math, argparse


def build(out_path):
    wb = openpyxl.Workbook()
    # --- 负对照:正常带噪声测量 ---
    ws = wb.active; ws.title = 'NEG_normal'
    ws.append(['x_axis', 'measure'])
    for i in range(40):
        x = round(i * 0.1, 1)
        y = round(2.0 + math.sin(i) * 1.7 + (i * 0.137 % 0.93), 3)   # 杂乱
        ws.append([x, y])
    # --- 1) 末位全是5 (1位小数) → B 末位锁定 ---
    ws = wb.create_sheet('PLANT_lastdigit5')
    ws.append(['groupA', 'groupB'])
    ints = [12, 7, 23, 4, 18, 9, 31, 6, 15, 28, 3, 19, 11, 26, 8, 14, 22, 5, 17, 29, 2, 13, 21, 33, 10, 16, 24, 1, 20, 30]
    for k in ints:
        ws.append([k + 0.5, (k * 1.3 % 40) + 0.5])           # 末位恒为5
    # --- 2) 两列恰好 +0.3 → A 两列恒差 ---
    ws = wb.create_sheet('PLANT_diff0.3')
    ws.append(['col3', 'col4'])
    base = [1.2, 3.7, 5.1, 8.9, 2.4, 6.6, 4.3, 7.8, 9.2, 0.5, 3.3, 5.9, 2.7, 6.1, 8.4, 1.8, 4.6, 7.2, 9.9, 0.8, 3.1, 5.4, 2.2, 6.8]
    for v in base:
        ws.append([round(v + 0.3, 1), round(v, 1)])          # 第1列 - 第2列 = 0.3
    # --- 3) 末两位全是44 → E 小数末两位雷同 ---
    ws = wb.create_sheet('PLANT_last2_44')
    ws.append(['vals'])
    for k in [1, 2, 7, 12, 5, 9, 3, 15, 8, 21, 6, 11, 4, 18, 10, 25, 13, 30, 2, 17]:
        c = ws.cell(ws.max_row + 1, 1, k + 0.44); c.number_format = '0.00'   # 末两位恒为44
    # --- 4) 同值霸占 (一个值反复) → D 同值霸占 ---
    ws = wb.create_sheet('PLANT_value_domination')
    ws.append(['x'])
    seq = [3.7] * 18 + [1.2, 4.5, 2.8, 5.1, 3.3, 6.6]
    for v in seq:
        ws.append([round(v, 2)])
    wb.save(out_path)
    print(f'written {out_path}')


if __name__ == '__main__':
    ap = argparse.ArgumentParser(description='Build a synthetic positive-control xlsx.')
    ap.add_argument('--out', default='./synthetic_positive.xlsx', help='output xlsx path')
    args = ap.parse_args()
    build(args.out)
